"""
Module contain external services with used in app
"""
from django.core import mail
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .models import Client
from app import settings


def get_clients_in_xlsx():
    """
    Write all clients data in xlsx document
    :return: xlsx workbook
    """
    clients = Client.objects.get_clients_data()
    wb = Workbook()
    ws = wb.active
    right_align = Alignment(horizontal='right')
    left_align = Alignment(horizontal='left')

    # Write header column
    ws.column_dimensions['A'].width = 20
    ws.cell(column=1, row=1, value='Client id')
    ws.cell(column=1, row=2, value='First name')
    ws.cell(column=1, row=3, value='Last name')
    ws.cell(column=1, row=4, value='Date of birth')
    ws.cell(column=1, row=5, value='Age')
    ws.cell(column=1, row=6, value='Photo').alignment = Alignment(vertical='center')

    # Write all client data
    col = 2
    for client in clients:
        ws.column_dimensions[get_column_letter(col)].width = 35

        # String client data
        date = try_parsing_date(client['date'])
        ws.cell(column=col, row=1, value=client['id']).alignment = right_align
        ws.cell(column=col, row=2, value=client['first_name']).alignment = right_align
        ws.cell(column=col, row=3, value=client['last_name']).alignment = right_align
        ws.cell(column=col, row=4, value=date).alignment = right_align
        ws.cell(column=col, row=5, value=client['age']).alignment = right_align

        # Image line
        ws.cell(column=col, row=6, value='You should see three logos below').alignment = left_align
        print(client['photo'].encode('utf-8'))
        img = Image('/code' + client['photo'])
        resize_image_to_xlsx(img)
        img_cell = get_column_letter(col) + '6'
        ws.add_image(img, img_cell)
        col += 1

    wb.save('client_data')
    wb.close()
    return wb


def resize_image_to_xlsx(img):
    """
    Resize image to base width with save proportions
    :param img: original size image
    """
    base_width = 252
    w_percent = (base_width / float(img.width))
    h_size = int((float(img.height) * float(w_percent)))
    img.width = base_width
    img.height = h_size


def try_parsing_date(date):
    """
    Get string by date, check all possible formats
    :param date: date
    :return: date in string format
    """
    for fmt in settings.DATE_INPUT_FORMATS:
        try:
            return date.strftime(fmt)
        except ValueError:
            pass
    raise ValueError('no valid date format found')
